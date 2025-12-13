# Eventlet monkey patching (최상단 필수!)
from __future__ import annotations
import eventlet
eventlet.monkey_patch()

from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_from_directory, send_file, Response
from flask_socketio import SocketIO, emit, join_room, leave_room
from flask_compress import Compress
from werkzeug.utils import secure_filename
from werkzeug.wrappers import Response as WerkzeugResponse
import json
import os
import uuid
import threading
from datetime import datetime
import logging
from logging.handlers import RotatingFileHandler
from typing import Any, Optional, Union
import database  # SQLite 데이터베이스 헬퍼
import pandas as pd
import random
from cache_manager import app_cache, cached, invalidate_cache, generate_etag
import push_helper  # 웹 푸시 알림 헬퍼
from rate_limiter import (
    create_limiter, get_limit_string, get_client_ip,
    check_login_lockout, record_login_attempt, get_remaining_attempts
)
from csrf_protection import init_csrf, exempt_csrf, is_csrf_exempt
from flasgger import Swagger
from swagger_config import SWAGGER_CONFIG, SWAGGER_TEMPLATE

# 로깅 설정
def setup_logging() -> logging.Logger:
    """애플리케이션 로깅 설정"""
    # 로그 디렉토리 생성
    log_dir = os.path.join(os.path.dirname(__file__), 'logs')
    os.makedirs(log_dir, exist_ok=True)

    # 로거 설정
    logger = logging.getLogger('crm')
    logger.setLevel(logging.INFO)

    # 파일 핸들러 (로테이션: 10MB, 최대 5개)
    file_handler = RotatingFileHandler(
        os.path.join(log_dir, 'crm.log'),
        maxBytes=10*1024*1024,
        backupCount=5,
        encoding='utf-8'
    )
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(logging.Formatter(
        '%(asctime)s - %(levelname)s - %(message)s'
    ))

    # 에러 전용 핸들러
    error_handler = RotatingFileHandler(
        os.path.join(log_dir, 'error.log'),
        maxBytes=10*1024*1024,
        backupCount=5,
        encoding='utf-8'
    )
    error_handler.setLevel(logging.ERROR)
    error_handler.setFormatter(logging.Formatter(
        '%(asctime)s - %(levelname)s - %(filename)s:%(lineno)d - %(message)s'
    ))

    logger.addHandler(file_handler)
    logger.addHandler(error_handler)

    return logger

logger = setup_logging()

app = Flask(__name__)
# Secret Key: 환경변수에서 로드, 없으면 기본값 (프로덕션에서는 반드시 환경변수 설정 필요)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'haru-crm-secret-key-2024-prod')
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB 최대 파일 크기
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 31536000  # 1년 캐시 (asset versioning으로 제어)
app.config['TEMPLATES_AUTO_RELOAD'] = True  # 템플릿 자동 리로드

# 세션 보안 설정
app.config['SESSION_COOKIE_SECURE'] = True  # HTTPS에서만 쿠키 전송
app.config['SESSION_COOKIE_HTTPONLY'] = True  # JavaScript에서 쿠키 접근 차단
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # CSRF 방지
app.config['PERMANENT_SESSION_LIFETIME'] = 3600  # 세션 타임아웃: 1시간 (초)

# Compression 설정
app.config['COMPRESS_MIMETYPES'] = [
    'text/html', 'text/css', 'text/javascript',
    'application/json', 'application/javascript'
]
app.config['COMPRESS_LEVEL'] = 6
app.config['COMPRESS_MIN_SIZE'] = 500

Compress(app)

# Rate Limiter 초기화
limiter = create_limiter(app)

# CSRF 보호 초기화
csrf = init_csrf(app)

# Swagger/OpenAPI 문서화 초기화
swagger = Swagger(app, config=SWAGGER_CONFIG, template=SWAGGER_TEMPLATE)

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf', 'doc', 'docx', 'xls', 'xlsx', 'txt', 'zip', 'rar', 'md', 'csv', 'json', 'xml', 'html', 'css', 'log'}
EXCEL_EXTENSIONS = {'xls', 'xlsx'}

# 파일 시그니처 (Magic Bytes) 검증
FILE_SIGNATURES = {
    'xlsx': [b'PK\x03\x04'],  # ZIP 기반 (Office Open XML)
    'xls': [b'\xd0\xcf\x11\xe0'],  # OLE2 Compound Document
    'png': [b'\x89PNG\r\n\x1a\n'],
    'jpg': [b'\xff\xd8\xff'],
    'jpeg': [b'\xff\xd8\xff'],
    'gif': [b'GIF87a', b'GIF89a'],
    'pdf': [b'%PDF'],
    'zip': [b'PK\x03\x04'],
    'rar': [b'Rar!\x1a\x07'],
}

def validate_file_signature(file_stream: Any, extension: str) -> bool:
    """파일 시그니처로 실제 파일 타입 검증"""
    ext = extension.lower()
    if ext not in FILE_SIGNATURES:
        return True  # 시그니처 정의가 없으면 통과

    file_stream.seek(0)
    header = file_stream.read(16)
    file_stream.seek(0)  # 스트림 위치 복원

    for sig in FILE_SIGNATURES[ext]:
        if header.startswith(sig):
            return True
    return False

# Socket.IO 초기화 (eventlet + Redis)
socketio = SocketIO(
    app,
    async_mode='eventlet',
    message_queue='redis://127.0.0.1:6379/0',
    cors_allowed_origins="*",
    max_http_buffer_size=50 * 1024 * 1024,
    logger=False,
    engineio_logger=False
)

# 관리자 사용자명 캐시 (DB에서 로드)
_admin_cache = None
_admin_cache_time = None

def get_admin_accounts() -> set[str]:
    """관리자 사용자명 집합 반환 (5분 캐시)"""
    global _admin_cache, _admin_cache_time
    import time
    now = time.time()

    # 캐시가 없거나 5분 지났으면 갱신
    if _admin_cache is None or _admin_cache_time is None or (now - _admin_cache_time) > 300:
        _admin_cache = database.get_admin_usernames()
        _admin_cache_time = now

    return _admin_cache

# SQLite 데이터베이스 함수 사용
load_data = database.load_data
save_data = database.save_data
load_chats = database.load_chats
save_chats = database.save_chats
load_users = database.load_users
save_users = database.save_users
load_promotions = database.load_promotions
save_promotions = database.save_promotions
add_user = database.add_user
load_users_by_team = database.load_users_by_team
load_teams = database.load_teams
load_users_with_team = database.load_users_with_team

def is_localhost() -> bool:
    """로컬호스트 여부 확인"""
    # Nginx 프록시 뒤에서는 X-Real-IP 헤더 사용
    real_ip = request.headers.get('X-Real-IP', request.remote_addr)
    return real_ip in ['127.0.0.1', 'localhost', '::1']

def is_admin() -> bool:
    """실제 관리자 권한 확인 (로컬호스트 또는 관리자 역할)"""
    if is_localhost():
        return True

    # 세션에 role이 있고 관리자인 경우
    if 'role' in session and session['role'] == '관리자':
        return True

    # DB에서 관리자 목록 확인
    if 'username' in session and session['username'] in get_admin_accounts():
        return True

    return False

def require_login() -> Optional[WerkzeugResponse]:
    """로그인 필수 체크 - 로그인하지 않은 경우 로그인 페이지로 리다이렉트"""
    if is_localhost():
        return None  # localhost는 항상 허용

    if 'username' not in session:
        return redirect(url_for('login'))

    return None

def require_admin() -> Optional[Union[WerkzeugResponse, str]]:
    """관리자 권한 필수 체크 - 권한 없으면 access_denied 페이지로"""
    if is_localhost():
        return None  # localhost는 항상 허용

    if 'username' not in session:
        return redirect(url_for('login'))

    if not is_admin():
        return render_template('access_denied.html',
                             message='이 페이지는 관리자만 접근할 수 있습니다.',
                             redirect_url=url_for('index'))

    return None

# Asset versioning for cache busting
_asset_manifest = None
def load_asset_manifest() -> dict[str, str]:
    """Load asset manifest with file hashes"""
    global _asset_manifest
    if _asset_manifest is None:
        manifest_path = os.path.join(app.root_path, 'static', 'asset_manifest.json')
        if os.path.exists(manifest_path):
            with open(manifest_path, 'r') as f:
                _asset_manifest = json.load(f)
        else:
            _asset_manifest = {}
    return _asset_manifest

@app.context_processor
def utility_processor():
    """템플릿에서 asset_version 함수 사용 가능하게"""
    def asset_version(filename):
        manifest = load_asset_manifest()
        # Extract base name without extension
        base_name = filename.rsplit('.', 1)[0] if '.' in filename else filename
        # Common.css → common_css 형식으로 변환
        key = base_name.replace('.', '_').replace('/', '_').replace('-', '_')

        if key in manifest:
            return f"{filename}?v={manifest[key]}"
        return filename

    return dict(asset_version=asset_version)

@app.before_request
def session_management():
    """세션 관리: 활동 갱신 및 보안 헤더"""
    # 정적 파일은 세션 체크 불필요
    if request.path.startswith('/static/') or request.path.startswith('/favicon'):
        return

    # 로그인된 사용자의 세션 활동 갱신
    if 'username' in session:
        session.modified = True  # 세션 타임아웃 갱신

@app.after_request
def add_security_headers(response):
    """보안 및 캐시 헤더 추가"""
    # 보안 헤더
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'

    # 정적 파일 (CSS, JS, 이미지, 폰트 등)은 1시간 캐싱
    if request.path.startswith('/static/') or request.path.startswith('/uploads/'):
        response.headers['Cache-Control'] = 'public, max-age=3600'
    # 동적 콘텐츠는 캐시 비활성화
    else:
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '-1'
    return response

@app.route('/')
def index():
    # localhost는 항상 admin으로
    if is_localhost():
        return redirect(url_for('admin'))

    # 세션 없으면 로그인 페이지로
    if 'username' not in session:
        return redirect(url_for('login'))

    # 관리자 계정은 관리자 페이지로
    if is_admin():
        return redirect(url_for('admin'))

    # 일반 사용자는 user 페이지로
    return render_template('user.html',
                         username=session['username'],
                         is_admin=False,
                         page_title='내 할일',
                         current_page='tasks')

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit(get_limit_string('login'))
def login():
    if is_localhost():
        return redirect(url_for('admin'))

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()

        if not username:
            return render_template('login.html', error='이름을 입력하세요.')

        if not password:
            return render_template('login.html', error='비밀번호를 입력하세요.', username=username)

        # 로그인 잠금 상태 확인
        locked, remaining = check_login_lockout(username)
        if locked:
            return render_template('login.html',
                error=f'너무 많은 로그인 시도로 {remaining}초간 잠겼습니다.',
                username=username,
                locked=True)

        # 데이터베이스에서 사용자 검증
        user = database.verify_user_login(username, password)

        if user:
            # 로그인 성공 - 시도 기록 초기화
            record_login_attempt(username, success=True)

            # 세션 재생성 (세션 고정 공격 방지)
            session.clear()
            session['username'] = user['username']
            session['role'] = user['role']
            session['login_time'] = datetime.now().isoformat()  # 로그인 시간 기록
            session.permanent = True  # 영구 세션 (PERMANENT_SESSION_LIFETIME 적용)

            logger.info(f"Login success: {username} from {get_client_ip()}")
            return redirect(url_for('index'))
        else:
            # 로그인 실패 - 시도 기록
            is_locked = record_login_attempt(username, success=False)
            remaining_attempts = get_remaining_attempts(username)
            logger.warning(f"Login failed: {username} from {get_client_ip()}, remaining: {remaining_attempts}")

            if is_locked:
                return render_template('login.html',
                    error='너무 많은 로그인 시도로 5분간 잠겼습니다.',
                    username=username,
                    locked=True)

            return render_template('login.html',
                error=f'이름 또는 비밀번호가 올바르지 않습니다. (남은 시도: {remaining_attempts}회)',
                username=username)

    return render_template('login.html')

@app.route('/admin')
def admin():
    # 관리자 권한 검증
    auth_check = require_admin()
    if auth_check:
        return auth_check

    username = session.get('username', 'Admin')
    return render_template('admin.html',
                         username=username,
                         is_admin=True,
                         page_title='할일 관리',
                         current_page='admin')

# 앱 버전 (배포 시 업데이트)
APP_VERSION = '2024121301'

@app.route('/api/version', methods=['GET'])
def get_version():
    """앱 버전 조회
    ---
    tags:
      - 시스템
    responses:
      200:
        description: 현재 앱 버전
        schema:
          type: object
          properties:
            version:
              type: string
              example: "2024121301"
    """
    return jsonify({'version': APP_VERSION})

@app.route('/api/items', methods=['GET'])
def get_items():
    """할일 목록 조회
    ---
    tags:
      - 할일(Task)
    security:
      - session: []
    responses:
      200:
        description: 할일 목록
        schema:
          type: array
          items:
            $ref: '#/definitions/Task'
      401:
        description: 인증 필요
        schema:
          $ref: '#/definitions/Error'
    """
    if is_admin():
        # 관리자는 전체 조회
        data = load_data()
        return jsonify(data)

    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    # 일반 사용자는 자신에게 할당된 항목만 조회 (최적화)
    username = session['username']
    user_items = database.load_data_by_assigned(username)
    return jsonify(user_items)

@app.route('/api/items', methods=['POST'])
def create_item():
    """할일 생성
    ---
    tags:
      - 할일(Task)
    security:
      - session: []
    parameters:
      - in: body
        name: body
        required: true
        schema:
          type: object
          required:
            - title
            - content
          properties:
            title:
              type: string
              description: 할일 제목
            content:
              type: string
              description: 할일 내용
            priority:
              type: string
              enum: [높음, 중간, 낮음]
              default: 중간
            due_date:
              type: string
              format: date
              description: 마감일
    responses:
      201:
        description: 생성된 할일
        schema:
          $ref: '#/definitions/Task'
      403:
        description: 권한 없음 (관리자 전용)
        schema:
          $ref: '#/definitions/Error'
    """
    if not is_admin():
        return jsonify({'error': 'Forbidden'}), 403

    data = load_data()
    new_item = request.json
    new_item['id'] = database.get_next_id('tasks')
    new_item['created_at'] = datetime.now().isoformat()
    data.append(new_item)
    save_data(data)

    return jsonify(new_item), 201

@app.route('/api/items/<int:item_id>', methods=['PUT'])
def update_item(item_id):
    """할일 수정
    ---
    tags:
      - 할일(Task)
    security:
      - session: []
    parameters:
      - in: path
        name: item_id
        type: integer
        required: true
        description: 할일 ID
      - in: body
        name: body
        required: true
        schema:
          type: object
          required:
            - title
            - content
          properties:
            title:
              type: string
            content:
              type: string
    responses:
      200:
        description: 수정된 할일
        schema:
          $ref: '#/definitions/Task'
      400:
        description: 잘못된 요청
      403:
        description: 권한 없음 (관리자 전용)
      404:
        description: 할일을 찾을 수 없음
    """
    if not is_admin():
        return jsonify({'error': 'Forbidden'}), 403

    updated_data = request.json
    title = updated_data.get('title')
    content = updated_data.get('content')

    if not title or not content:
        return jsonify({'error': 'Title and content required'}), 400

    success = database.update_task(item_id, title, content)

    if success:
        # 업데이트된 항목 반환
        data = database.load_data()
        for item in data:
            if item['id'] == item_id:
                return jsonify(item)
        return jsonify({'error': 'Not found'}), 404
    else:
        return jsonify({'error': 'Update failed'}), 500

@app.route('/api/items/<int:item_id>', methods=['DELETE'])
def delete_item(item_id):
    """할일 삭제
    ---
    tags:
      - 할일(Task)
    security:
      - session: []
    parameters:
      - in: path
        name: item_id
        type: integer
        required: true
        description: 삭제할 할일 ID
    responses:
      200:
        description: 삭제 결과
        schema:
          $ref: '#/definitions/Success'
      403:
        description: 권한 없음 (관리자 전용)
    """
    if not is_admin():
        return jsonify({'error': 'Forbidden'}), 403

    # 최적화된 delete_task 함수 사용 (개별 삭제)
    success = database.delete_task(item_id)

    return jsonify({'success': success})

@app.route('/api/items/<int:item_id>/unassign', methods=['POST'])
def unassign_item(item_id):
    """할일 배정 해제
    ---
    tags:
      - 할일(Task)
    security:
      - session: []
    parameters:
      - in: path
        name: item_id
        type: integer
        required: true
        description: 할일 ID
    responses:
      200:
        description: 배정 해제 성공
        schema:
          $ref: '#/definitions/Success'
      403:
        description: 권한 없음 (관리자 전용)
      500:
        description: 서버 오류
    """
    if not is_admin():
        return jsonify({'error': 'Forbidden'}), 403

    try:
        database.update_task_assignment(item_id, None)
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f'할일 배정 해제 실패 (item_id={item_id}): {e}', exc_info=True)
        return jsonify({'error': '배정 해제 중 오류가 발생했습니다'}), 500

@app.route('/api/items/<int:item_id>/status', methods=['PUT'])
def update_item_status(item_id):
    """할일 상태 변경
    ---
    tags:
      - 할일(Task)
    security:
      - session: []
    parameters:
      - in: path
        name: item_id
        type: integer
        required: true
        description: 할일 ID
      - in: body
        name: body
        required: true
        schema:
          type: object
          required:
            - status
          properties:
            status:
              type: string
              enum: [대기중, 진행중, 완료]
              description: 변경할 상태
    responses:
      200:
        description: 상태 변경 성공
        schema:
          type: object
          properties:
            success:
              type: boolean
            status:
              type: string
      400:
        description: 잘못된 상태값
      401:
        description: 인증 필요
      403:
        description: 권한 없음 (본인 할당 항목만)
      404:
        description: 할일을 찾을 수 없음
    """
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    status = request.json.get('status')
    if status not in ['대기중', '진행중', '완료']:
        return jsonify({'error': 'Invalid status'}), 400

    # 일반 사용자는 자신에게 배정된 항목만 변경 가능
    if not is_admin():
        # 최적화: 단일 항목만 조회
        with database.get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT assigned_to FROM tasks WHERE id = %s', (item_id,))
            row = cursor.fetchone()
            if not row:
                return jsonify({'error': 'Not found'}), 404
            if row['assigned_to'] != session['username']:
                return jsonify({'error': 'Forbidden'}), 403

    database.update_task_status(item_id, status)

    # Socket.IO로 할당자에게 배지 업데이트 전송
    with database.get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT assigned_to FROM tasks WHERE id = %s', (item_id,))
        row = cursor.fetchone()
        if row and row['assigned_to']:
            assignee = row['assigned_to']
            counts = calculate_nav_counts(assignee)
            socketio.emit('nav_counts_update', counts, room=f'user_{assignee}')

    return jsonify({'success': True, 'status': status})

@app.route('/api/items/<int:item_id>/assign', methods=['PUT'])
def update_item_assignment(item_id):
    """할일 배정/회수 (관리자 전용)"""
    if not is_admin():
        return jsonify({'error': 'Forbidden'}), 403

    assigned_to = request.json.get('assigned_to')  # None이면 회수

    # 이전 할당자 확인 (배지 업데이트용)
    old_assignee = None
    with database.get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT assigned_to FROM tasks WHERE id = %s', (item_id,))
        row = cursor.fetchone()
        if row:
            old_assignee = row['assigned_to']

    database.update_task_assignment(item_id, assigned_to)

    # Socket.IO로 배지 업데이트 전송 (이전 할당자 + 새 할당자)
    if old_assignee:
        counts = calculate_nav_counts(old_assignee)
        socketio.emit('nav_counts_update', counts, room=f'user_{old_assignee}')
    if assigned_to:
        counts = calculate_nav_counts(assigned_to)
        socketio.emit('nav_counts_update', counts, room=f'user_{assigned_to}')

    return jsonify({'success': True, 'assigned_to': assigned_to})

@app.route('/api/items/bulk-assign', methods=['POST'])
def bulk_assign_items():
    """할일 일괄 배정 (관리자 전용)"""
    if not is_admin():
        return jsonify({'error': 'Forbidden'}), 403

    task_ids = request.json.get('task_ids', [])
    assign_mode = request.json.get('mode')  # 'individual', 'random', 'sequential'
    users = request.json.get('users', [])  # 배정할 사용자 목록

    if not task_ids or not users:
        return jsonify({'error': 'Invalid parameters'}), 400

    try:
        if assign_mode == 'random':
            # 랜덤 배정:
            # 1. 먼저 항목 순서를 랜덤하게 섞음
            # 2. 나누어떨어지는 수만큼 균등 분배 (모두에게 최소 보장)
            # 3. 나머지는 랜덤하게 추가 분배
            shuffled_tasks = task_ids.copy()
            random.shuffle(shuffled_tasks)

            items_per_person = len(shuffled_tasks) // len(users)
            base_count = items_per_person * len(users)
            remainder = len(shuffled_tasks) - base_count

            logger.info(f"[랜덤배정] 전체: {len(shuffled_tasks)}개, 인원: {len(users)}명")
            logger.info(f"[랜덤배정] 1인당: {items_per_person}개, 균등배정: {base_count}개, 나머지: {remainder}개")

            # 1단계: 균등 분배 (모두에게 동일하게)
            for i, task_id in enumerate(shuffled_tasks[:base_count]):
                user = users[i % len(users)]
                database.update_task_assignment(task_id, user)

            logger.info(f"[랜덤배정] 1단계 완료: {base_count}개 균등 배정")

            # 2단계: 나머지를 랜덤하게 선정된 사람들에게 1개씩 분배
            remainder_tasks = shuffled_tasks[base_count:]
            logger.info(f"[랜덤배정] 2단계 시작: {len(remainder_tasks)}개 나머지 랜덤 배정")

            if remainder_tasks:
                # 나머지 개수만큼 사람을 랜덤하게 선정 (중복 없이)
                selected_users = random.sample(users, len(remainder_tasks))
                for task_id, user in zip(remainder_tasks, selected_users):
                    database.update_task_assignment(task_id, user)
                    logger.debug(f"[랜덤배정] 나머지 task_id={task_id} -> {user}")

            logger.info(f"[랜덤배정] 완료: 총 {len(shuffled_tasks)}개 배정")

        elif assign_mode == 'sequential':
            # 순차 배정: 딱 나누어떨어지는 수만큼만 순차 분배, 나머지는 미배정
            items_per_person = len(task_ids) // len(users)
            assignable_count = items_per_person * len(users)

            for i, task_id in enumerate(task_ids[:assignable_count]):
                user = users[i % len(users)]
                database.update_task_assignment(task_id, user)

        elif assign_mode == 'individual':
            # 개별 배정: task_ids와 users가 1:1 매칭
            for task_id, user in zip(task_ids, users):
                database.update_task_assignment(task_id, user)

        return jsonify({'success': True, 'count': len(task_ids)})

    except Exception as e:
        logger.error(f'일괄 배정 실패: {e}', exc_info=True)
        return jsonify({'error': '일괄 배정 중 오류가 발생했습니다'}), 500

@app.route('/api/items/bulk-upload', methods=['POST'])
@limiter.limit(get_limit_string('upload'))
def bulk_upload_items():
    """엑셀 파일로 할일 일괄 등록 (관리자 전용)"""
    if not is_admin():
        return jsonify({'error': 'Forbidden'}), 403

    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '파일이 선택되지 않았습니다'}), 400

    # 엑셀 파일 확인 (확장자)
    if not ('.' in file.filename and file.filename.rsplit('.', 1)[1].lower() in EXCEL_EXTENSIONS):
        return jsonify({'error': '엑셀 파일만 업로드 가능합니다 (.xlsx, .xls)'}), 400

    # 파일 시그니처 검증
    ext = file.filename.rsplit('.', 1)[1].lower()
    if not validate_file_signature(file.stream, ext):
        return jsonify({'error': '올바른 엑셀 파일이 아닙니다'}), 400

    try:
        # 엑셀 파일 읽기
        df = pd.read_excel(file)

        # 필수 컬럼 확인
        required_columns = ['제목', '내용']
        if not all(col in df.columns for col in required_columns):
            return jsonify({'error': f'필수 컬럼이 없습니다: {", ".join(required_columns)}'}), 400

        # 데이터 검증 및 등록
        added_count = 0
        skipped_count = 0

        for _, row in df.iterrows():
            title = str(row['제목']).strip() if pd.notna(row['제목']) else ''
            content = str(row['내용']).strip() if pd.notna(row['내용']) else ''

            # 제목이 비어있으면 스킵
            if not title:
                skipped_count += 1
                continue

            # 대상 처리 로직
            assigned_to = None
            if '대상' in df.columns and pd.notna(row['대상']):
                target_user = str(row['대상']).strip()
                if target_user:  # 빈 문자열이 아닌 경우만 검증
                    # DB에 사용자가 존재하는지 확인
                    if database.user_exists(target_user):
                        assigned_to = target_user
                    # 존재하지 않는 사용자명은 자동으로 미배정(None)으로 처리

            # 개별 삽입 (ID 자동 증가)
            database.add_task(assigned_to, title, content, '대기중')
            added_count += 1

        return jsonify({'success': True, 'count': added_count, 'skipped': skipped_count})

    except Exception as e:
        logger.error(f'엑셀 일괄 등록 실패: {e}', exc_info=True)
        return jsonify({'error': '파일 처리 중 오류가 발생했습니다'}), 500

@app.route('/api/users/non-admin', methods=['GET'])
def get_non_admin_users():
    """관리자가 아닌 일반 사용자 목록 조회 (관리자 전용)"""
    if not is_admin():
        return jsonify({'error': 'Forbidden'}), 403

    all_users = database.load_users()
    # 관리자 계정 제외
    admin_accounts = get_admin_accounts()
    non_admin_users = [user for user in all_users if user not in admin_accounts]
    return jsonify(non_admin_users)

@app.route('/api/teams', methods=['GET'])
def get_teams():
    """팀 목록 조회 (관리자 전용) - 배정 가능한 팀만 반환"""
    if not is_admin():
        return jsonify({'error': 'Forbidden'}), 403

    teams = load_teams()
    # '관리자' 팀은 배정 대상이 아니므로 제외
    teams = [team for team in teams if team != '관리자']
    return jsonify(teams)

@app.route('/api/users/by-team', methods=['GET'])
def get_users_by_team():
    """팀별 사용자 목록 조회 (관리자 전용)"""
    if not is_admin():
        return jsonify({'error': 'Forbidden'}), 403

    team = request.args.get('team')

    admin_accounts = get_admin_accounts()
    if team == '전체':
        # 전체 팀원 (관리자 팀 제외, 팀 없는 사용자 제외)
        users = load_users_by_team()
        # 관리자 계정은 배정 대상에서 제외
        users = [user for user in users if user not in admin_accounts]
    elif team == '관리자':
        # 관리자 팀은 배정 대상이 아니므로 빈 리스트 반환
        users = []
    elif team:
        # 특정 팀 사용자
        users = load_users_by_team(team)
        # 혹시 모를 경우를 대비해 관리자 제외
        users = [user for user in users if user not in admin_accounts]
    else:
        # 팀 파라미터 없으면 전체 사용자 (관리자 제외)
        users = database.load_all_users_detail()
        users = [user for user in users if user not in admin_accounts]

    return jsonify(users)

@app.route('/api/users/with-team', methods=['GET'])
def get_users_with_team():
    """팀 정보를 포함한 사용자 목록 조회 (관리자 전용)"""
    if not is_admin():
        return jsonify({'error': 'Forbidden'}), 403

    users = load_users_with_team()
    # 관리자 계정 제외
    admin_accounts = get_admin_accounts()
    users = [user for user in users if user['username'] not in admin_accounts]
    return jsonify(users)

@app.route('/download/template/tasks')
def download_tasks_template():
    """할일 등록용 엑셀 템플릿 다운로드"""
    if not is_admin():
        return "Access Denied", 403

    # 템플릿 데이터 생성
    template_data = {
        '제목': ['예시: 고객 문의 응대', '예시: 보고서 작성'],
        '내용': ['고객 A의 문의사항 확인 및 답변', '월간 실적 보고서 작성 및 제출'],
        '대상자': ['홍길동', '']  # 비워두면 미배정
    }

    df = pd.DataFrame(template_data)

    # 임시 파일로 저장
    template_path = os.path.join(app.config['UPLOAD_FOLDER'], 'tasks_template.xlsx')
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    df.to_excel(template_path, index=False, engine='openpyxl')

    return send_from_directory(app.config['UPLOAD_FOLDER'], 'tasks_template.xlsx',
                             as_attachment=True, download_name='할일목록_업로드양식.xlsx')

@app.route('/logout')
def logout():
    session.pop('username', None)
    return redirect(url_for('login'))

@app.route('/favicon.ico')
def favicon():
    # 빈 응답 반환 (404 오류 방지)
    return '', 204

# 파일 업로드/다운로드
def allowed_file(filename: str) -> bool:
    """허용된 파일 확장자인지 확인"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/upload', methods=['POST'])
@csrf.exempt  # 내부 AJAX API - 세션 인증으로 보호
@limiter.limit(get_limit_string('upload'))
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '파일이 선택되지 않았습니다'}), 400

    if file and allowed_file(file.filename):
        # 파일 시그니처 검증 (이미지/문서 파일)
        ext = file.filename.rsplit('.', 1)[1].lower()
        if ext in FILE_SIGNATURES and not validate_file_signature(file.stream, ext):
            return jsonify({'error': '올바른 파일 형식이 아닙니다'}), 400

        # 고유한 파일명 생성
        filename = secure_filename(file.filename)
        unique_filename = f"{uuid.uuid4().hex}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)

        # uploads 폴더가 없으면 생성
        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

        file.save(filepath)

        # 파일 정보 반환
        file_info = {
            'filename': filename,
            'unique_filename': unique_filename,
            'size': os.path.getsize(filepath),
            'url': f'/uploads/{unique_filename}'
        }

        return jsonify(file_info), 200

    return jsonify({'error': '허용되지 않는 파일 형식입니다'}), 400

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

# 채팅 관련 라우트
@app.route('/chats')
def chat_list():
    # 로그인 필수
    auth_check = require_login()
    if auth_check:
        return auth_check

    username = session.get('username', 'Admin')
    admin = is_admin()
    localhost = is_localhost()
    return render_template('chat_list.html',
                         username=username,
                         is_admin=admin,
                         is_localhost=localhost,
                         page_title='채팅 목록',
                         current_page='chats')

@app.route('/chats/all')
def chat_list_admin():
    """관리자 전용 채팅 관리 페이지 - /chats로 리다이렉트"""
    return redirect(url_for('chat_list'))

@app.route('/chat/create')
def chat_create_page():
    # 로그인 필수
    auth_check = require_login()
    if auth_check:
        return auth_check

    return render_template('chat_create.html')

@app.route('/chat/<chat_id>')
def chat_room(chat_id):
    # 로그인 필수
    auth_check = require_login()
    if auth_check:
        return auth_check

    chats = load_chats()
    if chat_id not in chats:
        return "Chat not found", 404

    username = session.get('username', 'Admin')
    chat_info = chats[chat_id]

    # 로컬호스트(진짜 서버 관리자)만 모든 채팅방 입장 가능
    # 관리자 계정 포함 일반 사용자는 자신이 참여자인 채팅방만 입장 가능
    if not is_localhost() and username not in chat_info['participants']:
        return render_template('access_denied.html',
                             message='이 채팅방의 참여자만 입장할 수 있습니다.',
                             redirect_url=url_for('chats'))

    return render_template('chat_room.html',
                         chat_id=chat_id,
                         username=username,
                         chat_title=chat_info['title'],
                         page_title=chat_info['title'],
                         current_page='chat_room',
                         is_admin=is_admin())

@app.route('/api/chats', methods=['GET'])
def get_chats():
    """채팅 목록 조회
    ---
    tags:
      - 채팅
    security:
      - session: []
    parameters:
      - in: query
        name: limit
        type: integer
        default: 1
        description: 각 채팅방당 반환할 메시지 개수
    responses:
      200:
        description: 채팅방 목록 (참여중인 채팅방만)
        schema:
          type: object
          additionalProperties:
            $ref: '#/definitions/Chat'
      401:
        description: 인증 필요
    """
    chats = load_chats()

    # limit 파라미터: 각 채팅방당 반환할 메시지 개수 (기본값: 1)
    message_limit = request.args.get('limit', 1, type=int)

    # 로컬호스트(진짜 서버 관리자)만 모든 채팅방 조회 가능
    if is_localhost():
        # 로컬호스트도 메시지 제한 적용
        optimized_chats = {}
        for chat_id, chat_info in chats.items():
            chat_data = chat_info.copy()
            if message_limit > 0 and 'messages' in chat_data:
                chat_data['messages'] = chat_data['messages'][-message_limit:]  # 최신 N개만
            optimized_chats[chat_id] = chat_data
        return jsonify(optimized_chats)

    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    # 관리자 계정 포함 모든 사용자는 자신이 참여한 채팅방만 조회
    username = session['username']
    user_chats = {}

    for chat_id, chat_info in chats.items():
        if username in chat_info['participants']:
            # 안 읽은 메시지 개수 계산
            unread_count = 0
            for msg in chat_info.get('messages', []):
                read_by = msg.get('read_by', [])
                # 내가 보낸 메시지가 아니고, 내가 읽지 않은 메시지
                if msg.get('username') != username and username not in read_by:
                    unread_count += 1

            # 디버깅 로그
            if unread_count > 0:
                logger.debug(f"[DEBUG] 채팅방 {chat_id} ({chat_info.get('title')}): 사용자 {username}의 안 읽은 메시지 = {unread_count}")

            # 채팅방 정보에 unread_count 추가 + 메시지 제한
            chat_data = chat_info.copy()
            chat_data['unread_count'] = unread_count

            # 최적화: 채팅 목록용으로는 최신 N개 메시지만 반환
            if message_limit > 0 and 'messages' in chat_data:
                chat_data['messages'] = chat_data['messages'][-message_limit:]

            user_chats[chat_id] = chat_data

    return jsonify(user_chats)

@app.route('/api/chats/all', methods=['GET'])
def get_all_chats():
    """관리자가 모든 채팅방 목록을 보기 위한 API (삭제 관리용)"""
    if not is_admin():
        return jsonify({'error': 'Forbidden'}), 403

    chats = load_chats()
    return jsonify(chats)

@app.route('/api/chats', methods=['POST'])
def create_chat():
    # 로그인한 모든 사용자가 채팅방 생성 가능
    if 'username' not in session and not is_localhost():
        return jsonify({'error': 'Unauthorized'}), 401

    chats = load_chats()
    new_chat = request.json

    # 다음 채팅방 ID 가져오기
    chat_id = str(database.get_next_id('chats'))

    creator = session.get('username', 'Admin')

    # 참여자 목록에 생성자 포함
    participants = new_chat.get('participants', [])
    if creator not in participants:
        participants.append(creator)

    # 1:1 채팅인지 확인 (참여자가 2명인 경우)
    is_one_to_one = len(participants) == 2

    # 1:1 채팅인 경우 상대방 이름을 제목으로 사용
    if is_one_to_one:
        other_user = [p for p in participants if p != creator][0]
        title = other_user
    else:
        # 다중 채팅인 경우 제목 필요
        title = new_chat.get('title', 'New Chat')

    chats[chat_id] = {
        'title': title,
        'participants': participants,
        'creator': creator,
        'messages': [],
        'is_one_to_one': is_one_to_one,
        'created_at': datetime.now().isoformat()
    }

    save_chats(chats)
    return jsonify({'chat_id': chat_id, 'chat': chats[chat_id]}), 201

@app.route('/api/chats/<chat_id>', methods=['DELETE'])
def delete_chat(chat_id):
    if not is_admin():
        return jsonify({'error': 'Forbidden'}), 403

    chats = load_chats()
    if chat_id in chats:
        del chats[chat_id]
        save_chats(chats)

    return jsonify({'success': True})

@app.route('/api/chats/<chat_id>/messages', methods=['GET'])
def get_chat_messages(chat_id):
    """
    채팅방 메시지 페이지네이션 API

    Query Parameters:
        - limit: 반환할 메시지 개수 (기본값: 50)
        - offset: 건너뛸 메시지 개수 (기본값: 0)
        - before_id: 특정 메시지 ID 이전의 메시지만 가져오기 (무한 스크롤용)

    Returns:
        {
            'messages': [...],
            'total': 전체 메시지 개수,
            'has_more': 더 가져올 메시지가 있는지 여부
        }
    """
    # 로그인 확인
    if 'username' not in session and not is_localhost():
        return jsonify({'error': 'Unauthorized'}), 401

    chats = load_chats()

    if chat_id not in chats:
        return jsonify({'error': 'Chat not found'}), 404

    chat_info = chats[chat_id]

    # 권한 확인: 참여자만 메시지 조회 가능
    if not is_localhost():
        username = session['username']
        if username not in chat_info['participants']:
            return jsonify({'error': 'Forbidden'}), 403

    # 파라미터 파싱
    limit = request.args.get('limit', 50, type=int)
    offset = request.args.get('offset', 0, type=int)
    before_id = request.args.get('before_id', type=int)  # 무한 스크롤용

    messages = chat_info.get('messages', [])
    total = len(messages)

    # before_id가 있으면 해당 메시지 이전의 메시지만 가져오기
    if before_id is not None:
        # 메시지에 ID가 없으면 인덱스를 ID로 사용
        before_index = None
        for idx, msg in enumerate(messages):
            msg_id = msg.get('id', idx)
            if msg_id == before_id:
                before_index = idx
                break

        if before_index is not None:
            # before_id 이전 메시지들만 선택
            messages = messages[:before_index]
            total = len(messages)

    # 오프셋과 리밋 적용 (최신 메시지부터 가져오려면 역순으로)
    start_idx = max(0, total - offset - limit)
    end_idx = total - offset

    paginated_messages = messages[start_idx:end_idx]
    has_more = start_idx > 0

    return jsonify({
        'messages': paginated_messages,
        'total': total,
        'has_more': has_more,
        'offset': offset,
        'limit': limit
    })

@app.route('/api/chats/<chat_id>/search', methods=['GET'])
@limiter.limit(get_limit_string('search'))
def search_chat_messages(chat_id):
    """
    채팅방 메시지 검색 API

    Query Parameters:
        - q: 검색어 (필수)
        - date: 특정 날짜 검색 (YYYY-MM-DD 형식)

    Returns:
        {
            'results': [{ message, index, highlight }...],
            'total': 검색 결과 개수
        }
    """
    if 'username' not in session and not is_localhost():
        return jsonify({'error': 'Unauthorized'}), 401

    chats = load_chats()

    if chat_id not in chats:
        return jsonify({'error': 'Chat not found'}), 404

    chat_info = chats[chat_id]

    # 권한 확인
    if not is_localhost():
        username = session['username']
        if username not in chat_info['participants']:
            return jsonify({'error': 'Forbidden'}), 403

    query = request.args.get('q', '').strip().lower()
    date_filter = request.args.get('date', '')  # YYYY-MM-DD 형식

    messages = chat_info.get('messages', [])
    results = []

    for idx, msg in enumerate(messages):
        msg_text = msg.get('message', '').lower()
        msg_timestamp = msg.get('timestamp', '')
        msg_date = msg_timestamp[:10] if msg_timestamp else ''  # YYYY-MM-DD 추출

        # 날짜 필터
        if date_filter and msg_date != date_filter:
            continue

        # 검색어 필터 (검색어가 있을 때만)
        if query and query not in msg_text:
            continue

        # 결과에 추가
        result_msg = msg.copy()
        result_msg['index'] = idx
        result_msg['id'] = msg.get('id', idx)
        results.append(result_msg)

    return jsonify({
        'results': results,
        'total': len(results),
        'query': query,
        'date': date_filter
    })

@app.route('/api/chats/<chat_id>/dates', methods=['GET'])
def get_chat_dates(chat_id):
    """
    채팅방에 메시지가 있는 날짜 목록 반환 (캘린더용)
    """
    if 'username' not in session and not is_localhost():
        return jsonify({'error': 'Unauthorized'}), 401

    chats = load_chats()

    if chat_id not in chats:
        return jsonify({'error': 'Chat not found'}), 404

    chat_info = chats[chat_id]

    # 권한 확인
    if not is_localhost():
        username = session['username']
        if username not in chat_info['participants']:
            return jsonify({'error': 'Forbidden'}), 403

    messages = chat_info.get('messages', [])
    dates = set()

    for msg in messages:
        timestamp = msg.get('timestamp', '')
        if timestamp:
            date = timestamp[:10]  # YYYY-MM-DD
            dates.add(date)

    return jsonify({
        'dates': sorted(list(dates)),
        'total': len(dates)
    })

@app.route('/api/chats/<chat_id>/messages/context/<int:msg_id>', methods=['GET'])
def get_message_context(chat_id, msg_id):
    """
    특정 메시지 ID를 기준으로 주변 메시지를 가져오는 API
    검색/날짜 이동 시 화면에 없는 메시지로 점프할 때 사용

    Query Parameters:
        - before: 해당 메시지 이전 몇 개를 가져올지 (기본: 25)
        - after: 해당 메시지 이후 몇 개를 가져올지 (기본: 25)

    Returns:
        {
            'messages': [...],
            'target_index': 타겟 메시지의 인덱스 (messages 배열 내),
            'first_msg_index': 전체 대화에서 첫 번째 메시지의 인덱스,
            'has_more_before': 더 이전 메시지가 있는지,
            'has_more_after': 더 이후 메시지가 있는지
        }
    """
    if 'username' not in session and not is_localhost():
        return jsonify({'error': 'Unauthorized'}), 401

    chats = load_chats()

    if chat_id not in chats:
        return jsonify({'error': 'Chat not found'}), 404

    chat_info = chats[chat_id]

    # 권한 확인
    if not is_localhost():
        username = session['username']
        if username not in chat_info['participants']:
            return jsonify({'error': 'Forbidden'}), 403

    before = int(request.args.get('before', 25))
    after = int(request.args.get('after', 25))

    messages = chat_info.get('messages', [])

    # 메시지 ID로 인덱스 찾기
    target_index = -1
    for idx, msg in enumerate(messages):
        if msg.get('id') == msg_id:
            target_index = idx
            break

    if target_index == -1:
        return jsonify({'error': 'Message not found'}), 404

    # 시작/끝 인덱스 계산
    start_index = max(0, target_index - before)
    end_index = min(len(messages), target_index + after + 1)

    # 메시지 슬라이스
    context_messages = messages[start_index:end_index]

    return jsonify({
        'messages': context_messages,
        'target_index': target_index - start_index,  # context_messages 내에서의 인덱스
        'first_msg_index': start_index,  # 전체 대화에서 첫 메시지 인덱스
        'total_messages': len(messages),
        'has_more_before': start_index > 0,
        'has_more_after': end_index < len(messages)
    })


@app.route('/api/search_users', methods=['GET'])
@limiter.limit(get_limit_string('search'))
def search_users():
    # 로그인한 사용자는 누구나 검색 가능
    if 'username' not in session and not is_localhost():
        return jsonify({'error': 'Unauthorized'}), 401

    query = request.args.get('q', '').lower()

    # 로그인한 사용자 목록 + 할당된 사용자 목록
    users = set(load_users())  # 로그인한 사용자들

    # 항목에 할당된 사용자도 추가
    data = load_data()
    for item in data:
        if item.get('assigned_to'):  # None이 아닌 경우만 추가
            users.add(item['assigned_to'])

    # 관리자 계정도 추가
    users.update(get_admin_accounts())

    # None 값 제거
    users.discard(None)

    if query:
        users = [u for u in users if u and query in u.lower()]
    else:
        users = [u for u in users if u]

    return jsonify(sorted(users))

# WebSocket 이벤트 핸들러
@socketio.on('join')
def on_join(data):
    chat_id = data['chat_id']
    username = data['username']
    join_room(chat_id)
    emit('user_joined', {'username': username}, room=chat_id)

@socketio.on('join_user_room')
def on_join_user_room(data):
    """사용자별 개인 room에 join (전역 알림용)"""
    username = data['username']
    join_room(f'user_{username}')

@socketio.on('leave')
def on_leave(data):
    chat_id = data['chat_id']
    username = data['username']
    leave_room(chat_id)
    emit('user_left', {'username': username}, room=chat_id)

@socketio.on('send_message')
def handle_message(data):
    chat_id = data['chat_id']
    username = data['username']
    message = data['message']
    file_info = data.get('file_info')  # 파일 정보

    # 메시지 객체 생성
    msg_obj = {
        'username': username,
        'message': message,
        'timestamp': datetime.now().isoformat(),
        'read_by': [username],  # 보낸 사람은 자동으로 읽음 처리
        'file_info': file_info  # 파일 정보 (이미지, 문서 등)
    }

    # file_info를 file_path/file_name으로 변환 (DB 저장용)
    if file_info and len(file_info) > 0:
        first_file = file_info[0]
        msg_obj['file_path'] = first_file.get('url')
        msg_obj['file_name'] = first_file.get('filename')

    # 최적화: 개별 메시지만 DB에 저장 (전체 로드/저장 제거)
    try:
        msg_id = database.save_message(chat_id, msg_obj)
        msg_obj['id'] = msg_id  # 클라이언트에 ID 전달
    except Exception as e:
        logger.error(f'메시지 저장 실패: {e}')
        return

    # 방의 모든 사용자에게 브로드캐스트
    emit('new_message', msg_obj, room=chat_id)

    # 채팅방 정보 조회 (최적화: 필요한 정보만 조회)
    chat_info = database.get_chat_info(chat_id)
    if not chat_info:
        return

    # 채팅 참여자들에게 알림 브로드캐스트
    for participant in chat_info['participants']:
        if participant != username:  # 보낸 사람 제외
            emit('global_new_message', {
                'chat_id': chat_id,
                'chat_title': chat_info['title'],
                'sender': username,
                'message': message[:100],
                'is_one_to_one': len(chat_info['participants']) == 2
            }, room=f'user_{participant}')

            # 캐시 무효화 후 네비게이션 배지 업데이트
            invalidate_cache(f'nav_counts:{participant}')
            participant_counts = calculate_nav_counts(participant)
            emit('nav_counts_update', participant_counts, room=f'user_{participant}')

            # 푸시 알림 발송 (백그라운드)
            target_user = participant  # 클로저 문제 방지
            def send_push_async(target=target_user):
                try:
                    push_helper.send_push_notification(
                        username=target,
                        title=f'{username}님의 메시지',
                        body=message[:100],
                        data={
                            'type': 'chat',
                            'chatId': chat_id,
                            'url': f'/chat/{chat_id}'
                        }
                    )
                except Exception as e:
                    logger.error(f'푸시 알림 발송 실패: {e}')

            threading.Thread(target=send_push_async).start()

@socketio.on('typing_start')
def handle_typing_start(data):
    chat_id = data['chat_id']
    username = data['username']
    # 같은 방의 다른 사용자들에게 타이핑 시작 알림
    emit('user_typing_start', {'username': username}, room=chat_id, include_self=False)

@socketio.on('typing_stop')
def handle_typing_stop(data):
    chat_id = data['chat_id']
    username = data['username']
    # 같은 방의 다른 사용자들에게 타이핑 중지 알림
    emit('user_typing_stop', {'username': username}, room=chat_id, include_self=False)

@socketio.on('mark_as_read')
def handle_mark_as_read(data):
    """메시지를 읽음으로 표시 (최적화: 직접 DB 업데이트 + 캐시 무효화)"""
    chat_id = data['chat_id']
    username = data['username']
    message_id = data.get('message_id')  # 특정 메시지 ID (옵션)

    try:
        if message_id is not None:
            # 특정 메시지만 읽음 처리
            database.mark_single_message_as_read(chat_id, message_id, username)
            # 읽음 상태 조회
            read_by = database.get_message_read_by(message_id)
            emit('read_receipt_update', {
                'chat_id': chat_id,
                'username': username,
                'message_id': message_id,
                'read_by': read_by
            }, room=chat_id, include_self=False)
        else:
            # 모든 메시지를 읽음 처리 (최적화된 함수 사용)
            affected = database.mark_messages_as_read(chat_id, username)
            logger.debug(f'채팅방 {chat_id}에서 {affected}개 메시지 읽음 처리 by {username}')

            emit('read_receipt_update', {
                'chat_id': chat_id,
                'username': username,
                'message_id': None,
                'all_messages_read': True
            }, room=chat_id, include_self=False)

        # 캐시 무효화 후 네비게이션 배지 업데이트
        invalidate_cache(f'nav_counts:{username}')
        user_counts = calculate_nav_counts(username)
        emit('nav_counts_update', user_counts, room=f'user_{username}')
    except Exception as e:
        logger.error(f'읽음 처리 실패 (chat_id={chat_id}): {e}')

# ============== 프로모션 게시판 ==============

# 프로모션 편집 잠금 상태 (promotion_id: username)
promotion_locks = {}

@socketio.on('lock_promotion_edit')
def handle_lock_promotion_edit(data):
    """프로모션 수정 잠금"""
    promo_id = str(data['promo_id'])
    username = data['username']

    # 이미 다른 사용자가 편집 중인지 확인
    if promo_id in promotion_locks and promotion_locks[promo_id] != username:
        emit('promotion_locked', {
            'promo_id': promo_id,
            'locked_by': promotion_locks[promo_id]
        })
        return

    # 잠금 설정
    promotion_locks[promo_id] = username

    # 모든 사용자에게 잠금 상태 브로드캐스트
    emit('promotion_edit_status', {
        'promo_id': promo_id,
        'locked_by': username,
        'is_locked': True
    }, broadcast=True)

@socketio.on('unlock_promotion_edit')
def handle_unlock_promotion_edit(data):
    """프로모션 수정 잠금 해제"""
    promo_id = str(data['promo_id'])
    username = data['username']

    # 본인이 잠근 경우만 해제 가능
    if promo_id in promotion_locks and promotion_locks[promo_id] == username:
        del promotion_locks[promo_id]

        # 모든 사용자에게 잠금 해제 브로드캐스트
        emit('promotion_edit_status', {
            'promo_id': promo_id,
            'locked_by': None,
            'is_locked': False
        }, broadcast=True)

@app.route('/promotions')
def promotions_page():
    # 로그인 필수
    auth_check = require_login()
    if auth_check:
        return auth_check

    return render_template('promotions.html',
                         username=session['username'],
                         is_admin=is_admin(),
                         page_title='프로모션 게시판',
                         current_page='promotions')

@app.route('/api/promotions', methods=['GET'])
def get_promotions():
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    promotions = load_promotions()

    # 필터링 (쿼리 파라미터)
    category = request.args.get('category')
    product_name = request.args.get('product_name')
    channel = request.args.get('channel')
    promotion_name = request.args.get('promotion_name')
    search = request.args.get('search')

    filtered = promotions

    if category:
        filtered = [p for p in filtered if p.get('category') == category]
    if product_name:
        filtered = [p for p in filtered if p.get('product_name') == product_name]
    if channel:
        filtered = [p for p in filtered if p.get('channel') == channel]
    if promotion_name:
        filtered = [p for p in filtered if p.get('promotion_name') == promotion_name]
    if search:
        # 검색어로 모든 필드 검색
        search_lower = search.lower()
        filtered = [p for p in filtered if
                   search_lower in str(p.get('product_name', '')).lower() or
                   search_lower in str(p.get('channel', '')).lower() or
                   search_lower in str(p.get('promotion_name', '')).lower() or
                   search_lower in str(p.get('promotion_code', '')).lower() or
                   search_lower in str(p.get('content', '')).lower()]

    return jsonify(filtered)

@app.route('/api/promotions', methods=['POST'])
def create_promotion():
    if not is_admin():
        return jsonify({'error': 'Forbidden'}), 403

    data = request.json

    # 필수 값 검증
    required_fields = ['category', 'product_name', 'channel', 'promotion_name', 'content', 'start_date']
    for field in required_fields:
        if not data.get(field):
            return jsonify({'error': f'{field}는 필수 항목입니다'}), 400

    promotions = load_promotions()

    new_promotion = {
        'id': database.get_next_id('promotions'),
        'category': data['category'],
        'product_name': data['product_name'],
        'channel': data['channel'],
        'promotion_name': data['promotion_name'],
        'discount_amount': data.get('discount_amount', ''),
        'session_exemption': data.get('session_exemption', ''),
        'subscription_types': data.get('subscription_types', []),
        'promotion_code': data.get('promotion_code', ''),
        'content': data['content'],
        'start_date': data['start_date'],
        'end_date': data.get('end_date', '무기한'),
        'created_at': datetime.now().isoformat(),
        'updated_at': datetime.now().isoformat(),
        'created_by': session['username']
    }

    promotions.append(new_promotion)
    save_promotions(promotions)

    return jsonify(new_promotion), 201

@app.route('/api/promotions/<int:promo_id>', methods=['GET'])
def get_promotion(promo_id):
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    promotions = load_promotions()
    promo = next((p for p in promotions if p.get('id') == promo_id), None)

    if not promo:
        return jsonify({'error': 'Not found'}), 404

    return jsonify(promo)

@app.route('/api/promotions/<int:promo_id>', methods=['PUT'])
def update_promotion(promo_id):
    if not is_admin():
        return jsonify({'error': 'Forbidden'}), 403

    data = request.json
    promotions = load_promotions()

    for i, p in enumerate(promotions):
        if p.get('id') == promo_id:
            # 필수 값 검증
            required_fields = ['category', 'product_name', 'channel', 'promotion_name', 'content', 'start_date']
            for field in required_fields:
                if field in data and not data[field]:
                    return jsonify({'error': f'{field}는 필수 항목입니다'}), 400

            # 업데이트
            promotions[i].update({
                'category': data.get('category', p.get('category', '안마의자')),
                'product_name': data.get('product_name', p['product_name']),
                'channel': data.get('channel', p['channel']),
                'promotion_name': data.get('promotion_name', p['promotion_name']),
                'discount_amount': data.get('discount_amount', p.get('discount_amount', '')),
                'session_exemption': data.get('session_exemption', p.get('session_exemption', '')),
                'subscription_types': data.get('subscription_types', p.get('subscription_types', [])),
                'promotion_code': data.get('promotion_code', p.get('promotion_code', '')),
                'content': data.get('content', p['content']),
                'start_date': data.get('start_date', p['start_date']),
                'end_date': data.get('end_date', p.get('end_date', '무기한')),
                'updated_at': datetime.now().isoformat()
            })

            save_promotions(promotions)
            return jsonify(promotions[i])

    return jsonify({'error': 'Not found'}), 404

@app.route('/api/promotions/<int:promo_id>', methods=['DELETE'])
def delete_promotion(promo_id):
    if not is_admin():
        return jsonify({'error': 'Forbidden'}), 403

    promotions = load_promotions()
    promotions = [p for p in promotions if p.get('id') != promo_id]
    save_promotions(promotions)

    return jsonify({'success': True})

@app.route('/api/promotions/filters', methods=['GET'])
def get_promotion_filters():
    """필터링을 위한 고유 값 목록 반환"""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    promotions = load_promotions()

    categories = list(set(p.get('category') for p in promotions if p.get('category')))
    products = list(set(p.get('product_name') for p in promotions if p.get('product_name')))
    channels = list(set(p.get('channel') for p in promotions if p.get('channel')))
    promo_names = list(set(p.get('promotion_name') for p in promotions if p.get('promotion_name')))

    # 대분류별 상품 매핑 생성
    category_products = {}
    for promo in promotions:
        cat = promo.get('category')
        prod = promo.get('product_name')
        if cat and prod:
            if cat not in category_products:
                category_products[cat] = set()
            category_products[cat].add(prod)

    # set을 list로 변환하고 정렬
    for cat in category_products:
        category_products[cat] = sorted(list(category_products[cat]))

    return jsonify({
        'categories': sorted(categories),
        'products': sorted(products),
        'channels': sorted(channels),
        'promotion_names': sorted(promo_names),
        'category_products': category_products  # 대분류별 상품 매핑 추가
    })

@app.route('/api/promotions/template', methods=['GET'])
def download_promotion_template():
    """프로모션 일괄등록 엑셀 양식 다운로드"""
    if not is_admin():
        return jsonify({'error': 'Forbidden'}), 403

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "프로모션 일괄등록"

    # 스타일 정의
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    required_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")  # 연한 노란색
    optional_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")  # 연한 회색
    example_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")  # 연한 파랑
    guide_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 연한 노랑

    header_font = Font(bold=True, color="FFFFFF", size=11)
    bold_font = Font(bold=True, size=10)
    normal_font = Font(size=10)
    small_font = Font(size=9, color="666666")

    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 제목 (1행)
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = "📊 프로모션 일괄등록 양식"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    title_cell.alignment = center_alignment
    ws.row_dimensions[1].height = 25

    # 필수/옵션 안내 (2행)
    ws.merge_cells('A2:K2')
    info_cell = ws['A2']
    info_cell.value = "💡 필수 항목(노란색 배경)은 반드시 입력해야 하며, 옵션 항목(회색 배경)은 선택사항입니다"
    info_cell.font = Font(size=10, color="C00000")
    info_cell.fill = guide_fill
    info_cell.alignment = left_alignment
    ws.row_dimensions[2].height = 30

    # 헤더 (3행) - 필수/옵션 표시
    headers_with_required = [
        ("대분류", True),       # 필수
        ("상품명", True),       # 필수
        ("채널", True),         # 필수
        ("프로모션명", True),   # 필수
        ("금액할인", False),    # 옵션
        ("회차면제", False),    # 옵션
        ("중복여부", False),    # 옵션
        ("프로모션코드", False),# 옵션
        ("프로모션내용", True), # 필수
        ("시작일", True),       # 필수
        ("종료일", False)       # 옵션
    ]

    for col_num, (header, is_required) in enumerate(headers_with_required, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = f"{header} *" if is_required else header
        cell.fill = required_fill if is_required else optional_fill
        cell.font = header_font if is_required else Font(bold=True, color="000000", size=11)
        cell.alignment = center_alignment
        cell.border = thin_border

    ws.row_dimensions[3].height = 35

    # 설명 (4행)
    descriptions = [
        "상품 분류",
        "등록된 상품명\n입력 필수",
        "판매 채널\n입력 필수",
        "프로모션 이름\n입력 필수",
        "할인 금액\n예: 10,000원",
        "면제 회차\n예: 1회차",
        "기존/결합/지인\n콤마로 구분",
        "프로모션 코드\n영문+숫자",
        "상세 설명\n입력 필수",
        "YYYY-MM-DD\n입력 필수",
        "YYYY-MM-DD\n비우면 무기한"
    ]

    for col_num, desc in enumerate(descriptions, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = desc
        cell.font = small_font
        cell.alignment = center_alignment
        cell.border = thin_border

    ws.row_dimensions[4].height = 40

    # 예시 데이터 1 (5행)
    example_data_1 = [
        "안마의자",
        "안마의자 프리미엄",
        "온라인",
        "신규고객 특별할인",
        "10,000원",
        "1회차 면제",
        "기존,결합",
        "NEW2024",
        "신규 가입 고객 대상 10,000원 할인 프로모션",
        "2024-01-01",
        "2024-12-31"
    ]

    for col_num, value in enumerate(example_data_1, 1):
        cell = ws.cell(row=5, column=col_num)
        cell.value = value
        cell.font = normal_font
        cell.fill = example_fill
        cell.alignment = left_alignment
        cell.border = thin_border

    ws.row_dimensions[5].height = 30

    # 예시 데이터 2 (6행)
    example_data_2 = [
        "공기청정기",
        "공기청정기 스탠다드",
        "오프라인",
        "회차면제 프로모션",
        "",
        "2회차 면제",
        "결합",
        "FREE2024",
        "첫 2회차 렌탈료 면제 프로모션",
        "2024-06-01",
        ""
    ]

    for col_num, value in enumerate(example_data_2, 1):
        cell = ws.cell(row=6, column=col_num)
        cell.value = value
        cell.font = normal_font
        cell.fill = example_fill
        cell.alignment = left_alignment
        cell.border = thin_border

    ws.row_dimensions[6].height = 30

    # 안내 메시지 (8-15행)
    ws.merge_cells('A8:K8')
    guide_title = ws['A8']
    guide_title.value = "📋 작성 가이드"
    guide_title.font = Font(bold=True, size=11, color="C00000")
    guide_title.fill = guide_fill
    guide_title.alignment = left_alignment

    # 9-16행: 가이드 (17행은 헤더이므로 여기서 멈춤)
    guides = [
        "",
        "💡 위 5-6행은 예시용입니다. 삭제하지 않아도 자동으로 무시됩니다.",
        "",
        "✅ 필수 항목 (노란색): 대분류, 상품명, 채널, 프로모션명, 프로모션내용, 시작일",
        "📌 옵션 항목 (회색): 금액할인, 회차면제, 중복여부, 프로모션코드, 종료일",
        "",
        "🔹 아래 17행은 헤더, 실제 데이터는 18행부터 입력하세요 🔹",
        ""
    ]

    for idx, guide in enumerate(guides, 9):
        ws.merge_cells(f'A{idx}:K{idx}')
        cell = ws[f'A{idx}']
        cell.value = guide
        cell.font = Font(size=9)
        cell.alignment = left_alignment
        if "⚠️" in guide or "🔹" in guide:
            cell.font = Font(bold=True, size=10, color="C00000")

    # 컬럼 너비 조정
    column_widths = [12, 20, 12, 20, 12, 12, 15, 15, 40, 12, 12]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col_num)].width = width

    # 데이터 입력 시작 행 (17행)에 헤더 재표시
    for col_num, (header, is_required) in enumerate(headers_with_required, 1):
        cell = ws.cell(row=17, column=col_num)
        cell.value = f"{header} *" if is_required else header
        cell.fill = required_fill if is_required else optional_fill
        cell.font = Font(bold=True, color="000000", size=10)
        cell.alignment = center_alignment
        cell.border = thin_border

    ws.row_dimensions[17].height = 25

    # 18행부터 여유 공간 (사용자 입력용)
    for row in range(18, 21):
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

    # BytesIO로 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='프로모션_일괄등록_양식.xlsx'
    )

@app.route('/api/promotions/bulk-upload', methods=['POST'])
@limiter.limit(get_limit_string('upload'))
def bulk_upload_promotions():
    """엑셀 파일을 파싱하여 JSON 형태로 반환 (저장하지 않음)"""
    if not is_admin():
        return jsonify({'error': 'Forbidden'}), 403

    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '파일이 선택되지 않았습니다'}), 400

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': '엑셀 파일만 업로드 가능합니다'}), 400

    # 파일 시그니처 검증
    ext = file.filename.rsplit('.', 1)[1].lower()
    if not validate_file_signature(file.stream, ext):
        return jsonify({'error': '올바른 엑셀 파일이 아닙니다'}), 400

    try:
        from openpyxl import load_workbook
        from io import BytesIO

        # 파일을 메모리에 로드
        file_content = BytesIO(file.read())
        wb = load_workbook(file_content, data_only=True)
        ws = wb.active

        # 헤더 확인 (17번째 행 - 데이터 입력 시작 행)
        expected_headers = [
            "대분류", "상품명", "채널", "프로모션명",
            "금액할인", "회차면제", "중복여부", "프로모션코드",
            "프로모션내용", "시작일", "종료일"
        ]

        # 헤더 검증 (17번째 행에서 헤더 확인, * 제거 후 비교)
        actual_headers = [str(cell.value).replace(' *', '').strip() if cell.value else '' for cell in ws[17]]
        if actual_headers[:len(expected_headers)] != expected_headers:
            return jsonify({'error': '엑셀 양식이 올바르지 않습니다. 제공된 양식을 사용해주세요.'}), 400

        # 데이터 파싱 (18번째 행부터, 예시 데이터와 가이드 제외하고 실제 데이터만)
        promotions_data = []
        errors = []

        for row_num, row in enumerate(ws.iter_rows(min_row=18, values_only=True), start=18):
            # 빈 행 스킵
            if not any(row):
                continue

            # 주의사항 행 스킵
            if row[0] and str(row[0]).startswith('※'):
                continue

            # 필수 필드 검증
            category, product_name, channel, promotion_name, \
            discount_amount, session_exemption, subscription_types_str, promotion_code, \
            content, start_date, end_date = row[:11]

            # 필수 필드 체크
            if not all([category, product_name, channel, promotion_name, content, start_date]):
                errors.append(f"행 {row_num}: 필수 항목이 누락되었습니다 (대분류, 상품명, 채널, 프로모션명, 내용, 시작일은 필수)")
                continue

            # 중복여부 파싱
            subscription_types = []
            if subscription_types_str:
                types_list = [t.strip() for t in str(subscription_types_str).split(',')]
                for t in types_list:
                    if t in ['기존', '결합', '지인']:
                        subscription_types.append(t)

            # 날짜 형식 변환
            def format_date(date_val):
                if date_val is None:
                    return None
                if isinstance(date_val, str):
                    return date_val
                # datetime 객체인 경우
                try:
                    return date_val.strftime('%Y-%m-%d')
                except:
                    return str(date_val)

            start_date_str = format_date(start_date)
            end_date_str = format_date(end_date) if end_date else '무기한'

            promotion_data = {
                'category': str(category) if category else '',
                'product_name': str(product_name),
                'channel': str(channel),
                'promotion_name': str(promotion_name),
                'discount_amount': str(discount_amount) if discount_amount else '',
                'session_exemption': str(session_exemption) if session_exemption else '',
                'subscription_types': subscription_types,
                'promotion_code': str(promotion_code) if promotion_code else '',
                'content': str(content),
                'start_date': start_date_str,
                'end_date': end_date_str
            }

            promotions_data.append(promotion_data)

        if errors:
            return jsonify({'error': '\n'.join(errors)}), 400

        if not promotions_data:
            return jsonify({'error': '등록할 데이터가 없습니다'}), 400

        return jsonify({
            'success': True,
            'count': len(promotions_data),
            'data': promotions_data
        })

    except Exception as e:
        return jsonify({'error': f'파일 처리 중 오류 발생: {str(e)}'}), 500

@app.route('/api/promotions/bulk-save', methods=['POST'])
def bulk_save_promotions():
    """미리보기에서 수정된 프로모션 목록을 일괄 저장"""
    if not is_admin():
        return jsonify({'error': 'Forbidden'}), 403

    try:
        data = request.get_json()
        promotions_to_save = data.get('promotions', [])

        if not promotions_to_save:
            return jsonify({'error': '저장할 데이터가 없습니다'}), 400

        # 기존 프로모션 로드
        existing_promotions = load_promotions()

        # 새 ID 생성을 위한 최대 ID 찾기
        max_id = max([p.get('id', 0) for p in existing_promotions], default=0)

        # 현재 사용자 정보
        username = session.get('username', 'Admin')
        now = datetime.now().isoformat()

        # 새 프로모션 추가
        for promo in promotions_to_save:
            max_id += 1
            promo['id'] = max_id
            promo['created_at'] = now
            promo['created_by'] = username
            promo['updated_at'] = now
            existing_promotions.append(promo)

        # 저장
        save_promotions(existing_promotions)

        return jsonify({
            'success': True,
            'count': len(promotions_to_save),
            'message': f'{len(promotions_to_save)}개의 프로모션이 등록되었습니다'
        })

    except Exception as e:
        return jsonify({'error': f'저장 중 오류 발생: {str(e)}'}), 500

@app.route('/api/promotions/bulk-update', methods=['PUT'])
def bulk_update_promotions():
    """선택된 프로모션들을 일괄 수정"""
    if not is_admin():
        return jsonify({'error': 'Forbidden'}), 403

    try:
        data = request.get_json()
        promotions_to_update = data.get('promotions', [])
        mode = data.get('mode', 'full')  # 'full' 또는 'quick'

        if not promotions_to_update:
            return jsonify({'error': '수정할 데이터가 없습니다'}), 400

        # 기존 프로모션 로드
        existing_promotions = load_promotions()
        promo_map = {p['id']: p for p in existing_promotions}

        # 현재 사용자 정보
        username = session.get('username', 'Admin')
        now = datetime.now().isoformat()

        updated_count = 0

        for update_data in promotions_to_update:
            promo_id = update_data.get('id')
            if promo_id not in promo_map:
                continue

            promo = promo_map[promo_id]

            if mode == 'full':
                # 전체 수정 모드: 모든 필드 업데이트
                for key in ['category', 'product_name', 'channel', 'promotion_name',
                           'discount_amount', 'session_exemption', 'subscription_types',
                           'promotion_code', 'content', 'start_date', 'end_date']:
                    if key in update_data:
                        promo[key] = update_data[key]
            else:
                # 빠른 수정 모드: 전달된 필드만 업데이트
                # 날짜 연장 처리
                if 'extend_days' in update_data:
                    extend_days = update_data['extend_days']
                    current_end_date = promo.get('end_date', '')
                    if current_end_date and current_end_date != '무기한':
                        try:
                            end_date = datetime.strptime(current_end_date, '%Y-%m-%d')
                            from datetime import timedelta
                            new_end_date = end_date + timedelta(days=extend_days)
                            promo['end_date'] = new_end_date.strftime('%Y-%m-%d')
                        except ValueError:
                            pass  # 날짜 파싱 실패 시 무시

                # 다른 필드 업데이트
                for key in ['start_date', 'end_date', 'channel', 'category']:
                    if key in update_data and key != 'extend_days':
                        promo[key] = update_data[key]

            promo['updated_at'] = now
            promo['updated_by'] = username
            updated_count += 1

        # 저장
        save_promotions(existing_promotions)

        return jsonify({
            'success': True,
            'count': updated_count,
            'message': f'{updated_count}개의 프로모션이 수정되었습니다'
        })

    except Exception as e:
        logger.error(f'일괄 수정 오류: {str(e)}')
        return jsonify({'error': f'수정 중 오류 발생: {str(e)}'}), 500

@app.route('/api/promotions/bulk-delete', methods=['DELETE'])
def bulk_delete_promotions():
    """선택된 프로모션들을 일괄 삭제"""
    if not is_admin():
        return jsonify({'error': 'Forbidden'}), 403

    try:
        data = request.get_json()
        ids_to_delete = set(data.get('ids', []))

        if not ids_to_delete:
            return jsonify({'error': '삭제할 프로모션이 없습니다'}), 400

        # 기존 프로모션 로드
        existing_promotions = load_promotions()

        # 삭제 대상 제외
        original_count = len(existing_promotions)
        remaining_promotions = [p for p in existing_promotions if p.get('id') not in ids_to_delete]
        deleted_count = original_count - len(remaining_promotions)

        # 저장
        save_promotions(remaining_promotions)

        return jsonify({
            'success': True,
            'count': deleted_count,
            'message': f'{deleted_count}개의 프로모션이 삭제되었습니다'
        })

    except Exception as e:
        logger.error(f'일괄 삭제 오류: {str(e)}')
        return jsonify({'error': f'삭제 중 오류 발생: {str(e)}'}), 500

# ==================== 개인 예약 관리 ====================

@app.route('/reminders')
def reminders_page():
    """예약 관리 페이지"""
    # 로그인 필수
    auth_check = require_login()
    if auth_check:
        return auth_check

    username = session.get('username', 'Admin')
    return render_template('reminders.html',
                         username=username,
                         is_admin=is_admin(),
                         page_title='내 예약 관리',
                         current_page='reminders')

# ==================== 마이페이지 ====================

@app.route('/mypage')
def mypage():
    """마이페이지"""
    # 로그인 필수
    auth_check = require_login()
    if auth_check:
        return auth_check

    username = session.get('username', 'Admin')
    user_info = database.get_user_info(username) if not is_localhost() else {
        'username': 'Admin',
        'role': '관리자',
        'team': None,
        'status': 'active',
        'created_at': None
    }

    return render_template('mypage.html',
                         username=username,
                         user_info=user_info,
                         is_admin=is_admin(),
                         page_title='마이페이지',
                         current_page='mypage')

@app.route('/api/change-password', methods=['POST'])
@limiter.limit("3 per minute")  # 비밀번호 변경은 엄격하게 제한
def change_password():
    """비밀번호 변경"""
    if 'username' not in session and not is_localhost():
        return jsonify({'error': 'Unauthorized'}), 401

    if is_localhost():
        return jsonify({'error': '로컬호스트에서는 비밀번호 변경이 불가능합니다.'}), 400

    data = request.json
    current_password = data.get('current_password', '').strip()
    new_password = data.get('new_password', '').strip()
    confirm_password = data.get('confirm_password', '').strip()

    if not current_password or not new_password or not confirm_password:
        return jsonify({'error': '모든 필드를 입력해주세요.'}), 400

    if new_password != confirm_password:
        return jsonify({'error': '새 비밀번호가 일치하지 않습니다.'}), 400

    if len(new_password) < 4:
        return jsonify({'error': '비밀번호는 최소 4자 이상이어야 합니다.'}), 400

    username = session.get('username')
    success, message = database.change_user_password(username, current_password, new_password)

    if success:
        return jsonify({'success': True, 'message': message})
    else:
        return jsonify({'error': message}), 400

@app.route('/api/reminders', methods=['GET'])
def get_reminders():
    """예약 목록 조회"""
    if 'username' not in session and not is_localhost():
        return jsonify({'error': 'Unauthorized'}), 401

    username = session.get('username', 'Admin')
    show_completed = request.args.get('show_completed', 'false').lower() == 'true'

    reminders = database.load_reminders(username, show_completed)
    return jsonify(reminders)

@app.route('/api/reminders', methods=['POST'])
def create_reminder():
    """새 예약 생성"""
    if 'username' not in session and not is_localhost():
        return jsonify({'error': 'Unauthorized'}), 401

    username = session.get('username', 'Admin')
    data = request.json

    title = data.get('title', '').strip()
    content = data.get('content', '').strip()
    scheduled_date = data.get('scheduled_date', '').strip()
    scheduled_time = data.get('scheduled_time', '').strip()

    if not title or not scheduled_date or not scheduled_time:
        return jsonify({'error': 'Missing required fields'}), 400

    reminder_id = database.add_reminder(username, title, content, scheduled_date, scheduled_time)

    # Socket.IO로 배지 업데이트 전송
    counts = calculate_nav_counts(username)
    socketio.emit('nav_counts_update', counts, room=f'user_{username}')

    return jsonify({'id': reminder_id, 'success': True}), 201

@app.route('/api/reminders/<int:reminder_id>', methods=['PUT'])
def update_reminder(reminder_id):
    """예약 수정"""
    if 'username' not in session and not is_localhost():
        return jsonify({'error': 'Unauthorized'}), 401

    username = session.get('username', 'Admin')
    data = request.json

    title = data.get('title', '').strip()
    content = data.get('content', '').strip()
    scheduled_date = data.get('scheduled_date', '').strip()
    scheduled_time = data.get('scheduled_time', '').strip()

    if not title or not scheduled_date or not scheduled_time:
        return jsonify({'error': 'Missing required fields'}), 400

    success = database.update_reminder(reminder_id, username, title, content, scheduled_date, scheduled_time)

    if not success:
        return jsonify({'error': 'Reminder not found or unauthorized'}), 404

    return jsonify({'success': True})

@app.route('/api/reminders/<int:reminder_id>', methods=['DELETE'])
def delete_reminder(reminder_id):
    """예약 삭제"""
    if 'username' not in session and not is_localhost():
        return jsonify({'error': 'Unauthorized'}), 401

    username = session.get('username', 'Admin')
    success = database.delete_reminder(reminder_id, username)

    if not success:
        return jsonify({'error': 'Reminder not found or unauthorized'}), 404

    return jsonify({'success': True})

@app.route('/api/reminders/<int:reminder_id>/complete', methods=['PATCH'])
def toggle_reminder_complete(reminder_id):
    """예약 완료 상태 토글"""
    if 'username' not in session and not is_localhost():
        return jsonify({'error': 'Unauthorized'}), 401

    username = session.get('username', 'Admin')
    success = database.toggle_reminder_complete(reminder_id, username)

    if not success:
        return jsonify({'error': 'Reminder not found or unauthorized'}), 404

    # Socket.IO로 배지 업데이트 전송
    counts = calculate_nav_counts(username)
    socketio.emit('nav_counts_update', counts, room=f'user_{username}')

    return jsonify({'success': True})

@app.route('/api/reminders/notifications', methods=['GET'])
def get_pending_notifications():
    """알림 필요한 예약 목록 (30분 전 알림용)"""
    if 'username' not in session and not is_localhost():
        return jsonify({'error': 'Unauthorized'}), 401

    username = session.get('username', 'Admin')
    reminders = database.get_pending_notifications(username)
    return jsonify(reminders)

@app.route('/api/reminders/<int:reminder_id>/notify', methods=['POST'])
def mark_reminder_notified(reminder_id):
    """30분 전 알림 발송 완료 표시"""
    if 'username' not in session and not is_localhost():
        return jsonify({'error': 'Unauthorized'}), 401

    database.mark_reminder_notified(reminder_id)
    return jsonify({'success': True})

@app.route('/api/reminders/banner-check', methods=['GET'])
def check_reminder_banner():
    """배너 표시용 예약 체크 (당일 + 지난 미완료 예약)"""
    if 'username' not in session and not is_localhost():
        return jsonify({'error': 'Unauthorized'}), 401

    username = session.get('username')
    from datetime import date
    today = str(date.today())

    # 당일 + 지난 미완료 예약 필터링
    today_count = 0
    overdue_count = 0

    # 로컬호스트 또는 로그인하지 않은 경우: 모든 사용자의 예약 확인
    if is_localhost() or not username:
        users = database.load_all_users_detail()
        for user in users:
            user_reminders = database.load_reminders(user['username'], show_completed=True)
            for r in user_reminders:
                # PostgreSQL: is_completed는 0(미완료) 또는 1(완료)
                if r.get('is_completed') == 1:
                    continue
                reminder_date = r.get('scheduled_date', '')
                if reminder_date == today:
                    today_count += 1
                elif reminder_date < today:
                    overdue_count += 1
    else:
        # 로그인된 사용자: 본인 예약만 확인
        reminders = database.load_reminders(username, show_completed=True)
        for r in reminders:
            # PostgreSQL: is_completed는 0(미완료) 또는 1(완료)
            if r.get('is_completed') == 1:
                continue
            reminder_date = r.get('scheduled_date', '')
            if reminder_date == today:
                today_count += 1
            elif reminder_date < today:
                overdue_count += 1

    return jsonify({
        'has_reminders': today_count > 0 or overdue_count > 0,
        'today_count': today_count,
        'overdue_count': overdue_count,
        'total_count': today_count + overdue_count
    })

@app.route('/api/reminders/today', methods=['GET'])
def get_today_reminders():
    """당일 예약 목록 (시간순 정렬)"""
    if 'username' not in session and not is_localhost():
        return jsonify({'error': 'Unauthorized'}), 401

    username = session.get('username')
    from datetime import date
    today = str(date.today())

    # 당일 미완료 예약 필터링
    today_reminders = []

    # 로컬호스트 또는 로그인하지 않은 경우: 모든 사용자의 예약 확인
    if is_localhost() or not username:
        users = database.load_all_users_detail()
        for user in users:
            user_reminders = database.load_reminders(user['username'])
            for r in user_reminders:
                if r.get('is_completed'):
                    continue
                if r.get('scheduled_date', '') == today:
                    today_reminders.append(r)
    else:
        # 로그인된 사용자: 본인 예약만 확인
        reminders = database.load_reminders(username)
        for r in reminders:
            if r.get('is_completed'):
                continue
            if r.get('scheduled_date', '') == today:
                today_reminders.append(r)

    # 시간순 정렬
    today_reminders.sort(key=lambda x: x.get('scheduled_time', ''))

    return jsonify(today_reminders)

@app.route('/api/holidays', methods=['GET'])
def get_holidays():
    """공휴일 목록 조회 (년도별)"""
    year = request.args.get('year', type=int)

    holidays = database.load_holidays(year)

    # 날짜를 키로 하는 딕셔너리로 변환
    holidays_dict = {}
    for h in holidays:
        date_str = str(h['holiday_date'])
        holidays_dict[date_str] = h['holiday_name']

    return jsonify(holidays_dict)

# ==================== 사용자 관리 API ====================

@app.route('/users')
def users_page():
    """사용자 관리 페이지 (관리자 전용)"""
    # 관리자 권한 검증
    auth_check = require_admin()
    if auth_check:
        return auth_check

    username = session.get('username', 'Admin')
    return render_template('users.html',
                         username=username,
                         is_admin=is_admin(),
                         page_title='사용자 관리',
                         current_page='users')

@app.route('/api/users', methods=['GET'])
def get_users():
    """사용자 목록 조회
    ---
    tags:
      - 사용자
    security:
      - session: []
    responses:
      200:
        description: 사용자 목록
        schema:
          type: array
          items:
            $ref: '#/definitions/User'
      401:
        description: 권한 없음 (관리자 전용)
    """
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401

    users = database.load_all_users_detail()
    return jsonify(users)

@app.route('/api/users', methods=['POST'])
def create_user_account():
    """사용자 생성
    ---
    tags:
      - 사용자
    security:
      - session: []
    parameters:
      - in: body
        name: body
        required: true
        schema:
          type: object
          required:
            - username
          properties:
            username:
              type: string
              description: 사용자명
            role:
              type: string
              enum: [관리자, 팀장, 상담사]
              default: 상담사
            status:
              type: string
              enum: [active, inactive]
              default: active
            team:
              type: string
              description: 소속 팀
    responses:
      200:
        description: 생성 성공
        schema:
          $ref: '#/definitions/Success'
      400:
        description: 잘못된 요청 또는 중복 사용자명
      401:
        description: 권한 없음 (관리자 전용)
    """
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401

    data = request.json
    username = data.get('username', '').strip()
    role = data.get('role', '상담사')
    status = data.get('status', 'active')
    team = data.get('team', '').strip() or None

    if not username or not role:
        return jsonify({'error': 'Username and role are required'}), 400

    # 초기 비밀번호 설정
    password = 'admin1234' if role == '관리자' else 'body123!'

    success = database.create_user(username, password, role, status, team)

    if success:
        return jsonify({'success': True, 'message': 'User created successfully'})
    else:
        return jsonify({'error': 'Username already exists'}), 400

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
def delete_user_account(user_id):
    """사용자 삭제 (관리자 전용)"""
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401

    success = database.delete_user(user_id)

    if success:
        return jsonify({'success': True, 'message': 'User deleted successfully'})
    else:
        return jsonify({'error': 'User not found'}), 404

@app.route('/api/users/<int:user_id>/status', methods=['PATCH'])
def update_user_status_api(user_id):
    """사용자 상태 변경 (관리자 전용)"""
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401

    data = request.json
    status = data.get('status')

    if status not in ['active', 'inactive']:
        return jsonify({'error': 'Invalid status'}), 400

    success = database.update_user_status(user_id, status)

    if success:
        return jsonify({'success': True, 'message': 'Status updated successfully'})
    else:
        return jsonify({'error': 'User not found'}), 404

@app.route('/api/users/<int:user_id>/team', methods=['PATCH'])
def update_user_team_api(user_id):
    """사용자 팀 변경 (관리자 전용)"""
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401

    data = request.json
    team_value = data.get('team')
    # None이거나 빈 문자열이면 None으로 처리
    team = team_value.strip() if team_value and isinstance(team_value, str) else None
    if team == '':
        team = None

    success = database.update_user_team(user_id, team)

    if success:
        return jsonify({'success': True, 'message': 'Team updated successfully'})
    else:
        return jsonify({'error': 'User not found'}), 404

@app.route('/api/users/<int:user_id>/role', methods=['PATCH'])
def update_user_role_api(user_id):
    """사용자 권한 변경 (관리자 전용)"""
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401

    data = request.json
    role = data.get('role')

    if role not in ['상담사', '관리자']:
        return jsonify({'error': 'Invalid role'}), 400

    success = database.update_user_role(user_id, role)

    if success:
        return jsonify({'success': True, 'message': 'Role updated successfully'})
    else:
        return jsonify({'error': 'User not found'}), 404

@app.route('/api/users/<int:user_id>/reset-password', methods=['POST'])
def reset_user_password_api(user_id):
    """사용자 비밀번호 초기화 (관리자 전용)"""
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401

    data = request.json
    role = data.get('role')

    if role not in ['상담사', '관리자']:
        return jsonify({'error': 'Invalid role'}), 400

    success = database.reset_user_password(user_id, role)

    if success:
        return jsonify({'success': True, 'message': 'Password reset successfully'})
    else:
        return jsonify({'error': 'User not found'}), 404

@app.route('/api/teams', methods=['GET'])
def get_teams_api():
    """팀 목록 조회
    ---
    tags:
      - 팀
    security:
      - session: []
    responses:
      200:
        description: 팀 목록
        schema:
          type: array
          items:
            $ref: '#/definitions/Team'
      401:
        description: 권한 없음 (관리자 전용)
    """
    if not is_admin():
        return jsonify({'error': 'Unauthorized'}), 401

    teams = database.load_teams()
    return jsonify(teams)

# ==================== 푸시 알림 API ====================

@app.route('/api/push/vapid-public-key', methods=['GET'])
def get_vapid_public_key():
    """VAPID 공개키 조회
    ---
    tags:
      - 푸시알림
    responses:
      200:
        description: VAPID 공개키 (Base64 인코딩)
        schema:
          type: object
          properties:
            publicKey:
              type: string
              description: Base64 URL-safe 인코딩된 공개키
      500:
        description: 서버 오류
    """
    try:
        vapid_keys = push_helper.get_vapid_keys()
        from py_vapid import Vapid
        from cryptography.hazmat.primitives import serialization
        import base64

        vapid = Vapid.from_file('/svc/was/crm/crm-webapp/vapid_private.pem')
        public_key_bytes = vapid.public_key.public_bytes(
            encoding=serialization.Encoding.X962,
            format=serialization.PublicFormat.UncompressedPoint
        )
        public_key_b64 = base64.urlsafe_b64encode(public_key_bytes).decode('utf-8').rstrip('=')

        return jsonify({'publicKey': public_key_b64})
    except Exception as e:
        logger.error(f'VAPID 공개키 조회 실패: {e}', exc_info=True)
        return jsonify({'error': '공개키 조회 중 오류가 발생했습니다'}), 500

@app.route('/api/push/subscribe', methods=['POST'])
@csrf.exempt  # 내부 AJAX API - 세션 인증으로 보호
def push_subscribe():
    """푸시 알림 구독
    ---
    tags:
      - 푸시알림
    security:
      - session: []
    parameters:
      - in: body
        name: body
        required: true
        schema:
          type: object
          required:
            - subscription
          properties:
            subscription:
              type: object
              description: PushSubscription 객체
              properties:
                endpoint:
                  type: string
                keys:
                  type: object
                  properties:
                    p256dh:
                      type: string
                    auth:
                      type: string
    responses:
      200:
        description: 구독 성공
        schema:
          $ref: '#/definitions/Success'
      400:
        description: 잘못된 요청
      401:
        description: 인증 필요
    """
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    try:
        data = request.json
        username = session.get('username')
        logger.debug(f"[Push Subscribe] username: {username}")

        if not data or 'subscription' not in data:
            return jsonify({'error': 'Invalid subscription data'}), 400

        subscription = data['subscription']

        if push_helper.save_subscription(username, subscription):
            logger.info(f"[Push Subscribe] 저장 성공: {username}")
            return jsonify({'success': True, 'message': 'Subscription saved'})
        else:
            logger.warning(f"[Push Subscribe] 저장 실패: {username}")
            return jsonify({'error': 'Failed to save subscription'}), 500

    except Exception as e:
        logger.error(f"[Push Subscribe] 예외 발생: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/api/push/unsubscribe', methods=['POST'])
def push_unsubscribe():
    """푸시 알림 구독 취소"""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    try:
        data = request.json
        endpoint = data.get('endpoint')

        if not endpoint:
            return jsonify({'error': 'Endpoint is required'}), 400

        if push_helper.remove_subscription(endpoint):
            return jsonify({'success': True, 'message': 'Subscription removed'})
        else:
            return jsonify({'error': 'Failed to remove subscription'}), 500

    except Exception as e:
        logger.error(f'푸시 구독 해제 실패: {e}', exc_info=True)
        return jsonify({'error': '구독 해제 중 오류가 발생했습니다'}), 500

@cached(ttl=10, key_prefix='nav_counts')
def calculate_nav_counts(username: str) -> dict[str, int]:
    """네비게이션 바 카운트 계산 (최적화: 전용 쿼리 사용, 10초 캐시)"""
    counts = {
        'pending_tasks': 0,
        'unread_chats': 0,
        'today_reminders': 0
    }

    try:
        # 읽지 않은 채팅 메시지 개수 (최적화: 전용 카운트 쿼리)
        counts['unread_chats'] = database.get_unread_chat_count(username)

        # 상담사: 내게 할당된 미완료 할일 개수 (assigned_to 사용)
        if username not in get_admin_accounts():
            with database.get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT COUNT(*) as count
                    FROM tasks
                    WHERE assigned_to = %s AND status != '완료'
                ''', (username,))
                row = cursor.fetchone()
                counts['pending_tasks'] = row['count'] if row else 0

        # 당일 예약 개수
        from datetime import date
        today = str(date.today())
        reminders = database.load_reminders(username)
        today_reminders = [r for r in reminders if r.get('scheduled_date') == today and not r.get('completed')]
        counts['today_reminders'] = len(today_reminders)

    except Exception as e:
        logger.error(f"Error calculating nav counts for {username}: {e}")

    return counts

@app.route('/api/nav-counts', methods=['GET'])
def get_nav_counts():
    """네비게이션 카운트 조회
    ---
    tags:
      - 시스템
    security:
      - session: []
    responses:
      200:
        description: 네비게이션 바 배지 카운트
        schema:
          $ref: '#/definitions/NavCounts'
      401:
        description: 인증 필요
    """
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    username = session['username']
    counts = calculate_nav_counts(username)
    logger.debug(f"[DEBUG] Final nav-counts for {username}: {counts}")
    return jsonify(counts)

# ============================================================
# 커스텀 에러 핸들러 (nginx 에러 페이지 사용)
# ============================================================
@app.errorhandler(400)
def bad_request(error):
    return send_file('/svc/web/nginx/html/errors/400.html'), 400

@app.errorhandler(401)
def unauthorized(error):
    return send_file('/svc/web/nginx/html/errors/401.html'), 401

@app.errorhandler(403)
def forbidden(error):
    return send_file('/svc/web/nginx/html/errors/403.html'), 403

@app.errorhandler(404)
def not_found(error):
    return send_file('/svc/web/nginx/html/errors/404.html'), 404

@app.errorhandler(500)
def internal_server_error(error):
    import traceback
    logger.error(f"500 Error: {request.path} - {error}\n{traceback.format_exc()}")
    return send_file('/svc/web/nginx/html/errors/500.html'), 500

@app.errorhandler(Exception)
def handle_exception(error):
    """처리되지 않은 모든 예외 로깅"""
    import traceback
    logger.error(f"Unhandled Exception: {request.path} - {type(error).__name__}: {error}\n{traceback.format_exc()}")
    return send_file('/svc/web/nginx/html/errors/500.html'), 500

@app.errorhandler(502)
def bad_gateway(error):
    return send_file('/svc/web/nginx/html/errors/502.html'), 502

@app.errorhandler(503)
def service_unavailable(error):
    return send_file('/svc/web/nginx/html/errors/503.html'), 503

if __name__ == '__main__':
    socketio.run(app, host='0.0.0.0', port=5000, debug=False, allow_unsafe_werkzeug=True)


# ==================== 알림 설정 API ====================

@app.route('/api/notification-settings', methods=['GET'])
def get_notification_settings():
    """알림 설정 조회
    ---
    tags:
      - 푸시알림
    security:
      - session: []
    responses:
      200:
        description: 사용자 알림 설정
        schema:
          $ref: '#/definitions/NotificationSettings'
      401:
        description: 인증 필요
    """
    if 'username' not in session and not is_localhost():
        return jsonify({'error': 'Unauthorized'}), 401

    username = session.get('username', 'Admin')
    settings = database.get_user_notification_settings(username)
    return jsonify(settings)


@app.route('/api/notification-settings', methods=['POST'])
def save_notification_settings():
    """사용자 알림 설정 저장"""
    if 'username' not in session and not is_localhost():
        return jsonify({'error': 'Unauthorized'}), 401

    username = session.get('username', 'Admin')
    data = request.get_json()

    settings = {
        'reminder_minutes': data.get('reminder_minutes', 30),
        'repeat_enabled': data.get('repeat_enabled', False),
        'repeat_interval': data.get('repeat_interval', 5),
        'repeat_until_minutes': data.get('repeat_until_minutes', 0),
        'daily_summary_enabled': data.get('daily_summary_enabled', True),
        'daily_summary_time': data.get('daily_summary_time', '09:00')
    }

    success = database.save_user_notification_settings(username, settings)

    if success:
        return jsonify({'success': True, 'settings': settings})
    else:
        return jsonify({'error': 'Failed to save settings'}), 500


@app.route('/api/notification-settings/test-daily-summary', methods=['POST'])
def test_daily_summary():
    """일일 요약 알림 테스트"""
    if 'username' not in session and not is_localhost():
        return jsonify({'error': 'Unauthorized'}), 401

    username = session.get('username', 'Admin')
    today_count = database.get_today_reminder_count(username)

    if today_count > 0:
        today_reminders = database.get_today_reminders_list(username)
        reminder_list = ', '.join([f"{r['scheduled_time']} {r['title']}" for r in today_reminders[:3]])
        if today_count > 3:
            reminder_list += f' 외 {today_count - 3}건'

        result = push_helper.send_push_notification(
            username=username,
            title='📅 오늘의 예약 알림 (테스트)',
            body=f'오늘 {today_count}건의 예약이 있습니다: {reminder_list}',
            data={
                'type': 'daily_summary',
                'url': '/reminders',
                'requireInteraction': False,
                'tag': 'daily-summary-test'
            }
        )
        return jsonify(result)
    else:
        return jsonify({'message': '오늘 예약이 없습니다', 'success': 0, 'failed': 0})


@app.route('/api/push/test', methods=['POST'])
def test_push_notification():
    """푸시 알림 테스트 API"""
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    username = session.get('username')

    try:
        result = push_helper.send_push_notification(
            username=username,
            title='테스트 푸시 알림',
            body='푸시 알림이 정상 작동합니다!',
            data={
                'type': 'test',
                'timestamp': str(datetime.now())
            }
        )

        logger.info(f'[PUSH TEST] 결과: {result}')
        return jsonify(result)
    except Exception as e:
        logger.error(f'[PUSH TEST] 오류: {e}')
        return jsonify({'error': str(e)}), 500

@app.route('/api/sw-version', methods=['GET'])
def get_sw_version():
    """Service Worker 버전 정보 조회 (클라이언트 자동 업데이트용)"""
    # Service Worker 파일의 수정 시간을 기반으로 타임스탬프 생성
    sw_path = os.path.join(os.path.dirname(__file__), 'static', 'service-worker.js')
    try:
        mtime = os.path.getmtime(sw_path)
        version = {
            'version': 'v9',  # 주 버전 (수동 관리)
            'timestamp': int(mtime),  # 파일 수정 시간
            'hash': str(int(mtime))  # 간단한 해시
        }
        return jsonify(version)
    except Exception as e:
        return jsonify({'version': 'v9', 'timestamp': 0, 'hash': '0'})


@app.route('/service-worker.js')
def serve_service_worker():
    """루트에서 Service Worker 제공 (scope='/'를 위해 필요)"""
    from flask import send_file
    sw_path = os.path.join(os.path.dirname(__file__), 'static', 'service-worker.js')
    response = send_file(sw_path, mimetype='application/javascript')
    # Service-Worker-Allowed 헤더로 루트 스코프 허용
    response.headers['Service-Worker-Allowed'] = '/'
    return response


# ==================== 예약 알림 스케줄러 ====================

def check_reminder_notifications():
    """예약 알림 체크 및 푸시 발송 (사용자별 설정 적용, 반복 알림 지원)"""
    while True:
        try:
            # 사용자별 설정이 적용된 알림 대상 목록 가져오기
            pending_reminders = database.get_pending_reminders_for_notification()

            for reminder in pending_reminders:
                if not reminder.get('should_notify'):
                    continue

                user_id = reminder.get('user_id')
                reminder_id = reminder.get('id')
                title = reminder.get('title', '예약')
                scheduled_time = reminder.get('scheduled_time', '')
                notify_reason = reminder.get('notify_reason', 'first')

                # 반복 알림인 경우 메시지 다르게
                if notify_reason == 'repeat':
                    notification_count = reminder.get('notification_count', 0)
                    push_title = f'🔔 재알림: {title}'
                    push_body = f'{scheduled_time} 예약이 곧 시작됩니다! (알림 {notification_count + 1}회차)'
                else:
                    push_title = f'⏰ 예약 알림: {title}'
                    push_body = f'{scheduled_time}에 "{title}" 예약이 있습니다.'

                # 푸시 알림 발송
                push_result = push_helper.send_push_notification(
                    username=user_id,
                    title=push_title,
                    body=push_body,
                    data={
                        'type': 'reminder',
                        'reminderId': reminder_id,
                        'url': '/reminders',
                        'requireInteraction': True,
                        'tag': f'reminder-{reminder_id}-{notify_reason}'
                    }
                )

                # 알림 발송 완료 표시
                if push_result.get('success', 0) > 0:
                    database.update_reminder_notification(reminder_id)
                    logger.info(f"[Reminder] 푸시 알림 발송 완료: {user_id} - {title} ({notify_reason})")
                else:
                    logger.warning(f"[Reminder] 푸시 알림 발송 실패: {user_id} - {title}, errors: {push_result.get('errors')}")

        except Exception as e:
            logger.error(f"[Reminder] 알림 체크 오류: {e}", exc_info=True)

        # 1분마다 체크
        eventlet.sleep(60)


def check_daily_summary_notifications():
    """아침 일일 요약 알림 체크 및 발송"""
    from datetime import date

    while True:
        try:
            now = datetime.now()
            current_time = now.strftime('%H:%M')
            today = str(date.today())

            # 일일 요약이 필요한 사용자 목록
            users_needing_summary = database.get_users_needing_daily_summary()

            for user in users_needing_summary:
                username = user.get('username')
                summary_time = user.get('daily_summary_time', '09:00')

                # 설정된 시간이 지났으면 알림 발송
                if current_time >= summary_time:
                    # 오늘 예약 개수 확인
                    today_count = database.get_today_reminder_count(username)

                    if today_count > 0:
                        # 오늘 예약 목록 가져오기
                        today_reminders = database.get_today_reminders_list(username)
                        reminder_list = ', '.join([f"{r['scheduled_time']} {r['title']}" for r in today_reminders[:3]])
                        if today_count > 3:
                            reminder_list += f' 외 {today_count - 3}건'

                        push_result = push_helper.send_push_notification(
                            username=username,
                            title=f'📅 오늘의 예약 알림',
                            body=f'오늘 {today_count}건의 예약이 있습니다: {reminder_list}',
                            data={
                                'type': 'daily_summary',
                                'url': '/reminders',
                                'requireInteraction': False,
                                'tag': f'daily-summary-{today}'
                            }
                        )

                        if push_result.get('success', 0) > 0:
                            database.update_last_daily_summary(username, today)
                            logger.info(f"[DailySummary] 일일 요약 발송 완료: {username} - {today_count}건")
                    else:
                        # 예약이 없어도 발송 완료 처리 (다시 보내지 않도록)
                        database.update_last_daily_summary(username, today)

        except Exception as e:
            logger.error(f"[DailySummary] 일일 요약 체크 오류: {e}", exc_info=True)

        # 5분마다 체크
        eventlet.sleep(300)


def start_reminder_scheduler():
    """예약 알림 스케줄러 시작"""
    logger.info("[Reminder] 예약 알림 스케줄러 시작")
    eventlet.spawn(check_reminder_notifications)
    eventlet.spawn(check_daily_summary_notifications)


# 앱 시작 시 스케줄러 실행 (Gunicorn 워커당 1회)
_scheduler_started = False

@app.before_request
def ensure_scheduler_started():
    """첫 요청 시 스케줄러 시작 (1회만)"""
    global _scheduler_started
    if not _scheduler_started:
        _scheduler_started = True
        start_reminder_scheduler()
